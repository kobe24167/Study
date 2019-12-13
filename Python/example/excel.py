"""
主要功能：
1、将需求的字段生成对应表格；2、将金额为正的（S）的行项目标记为黄色；
3、只导入特定物料；4、对数据源取名进行约束；
5、增加程序运行完毕的消息提醒；6、增加padas的排序、筛选功能
"""
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Border,Side,Alignment,Protection,Font,GradientFill,Color,colors
thin = Side(border_style="thin", color="000000")   #边框样式，定义为对象
align = Alignment(horizontal='center',vertical='center',wrap_text=True)
wb = load_workbook(filename=u"C:\\Users\\wbliang\\Desktop\\deptbudget1203.xlsx")
ws=wb["ExportExcel"]
for col in ["I","J"]:
    ws.column_dimensions[col].width=25   #设置固定列宽
for col in range(1,11):
    ws.cell(1,col).font=Font(name='Microsoft YaHei',size=13,bold=True,color=colors.RED)    #设置标题行的字体样式
    ws.cell(col,10).border=Border(top=thin, left=thin, right=thin, bottom=thin)   #设置标题行的边框样式
    ws.cell(col,10).alignment = align
wb.save("C:\\Users\\wbliang\\Desktop\\deptbudget1203.xlsx")
